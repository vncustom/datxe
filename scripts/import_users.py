"""Đọc file danh sách người dùng (.xlsx) -> prisma/data/users.json (đã chuẩn hoá).

Dùng lại mỗi khi cập nhật file Excel:
    python scripts/import_users.py "C:/Users/Admin/Downloads/Danh sách_user.xlsx"
"""
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = r"C:\Users\Admin\Downloads\Danh sách_user.xlsx"
OUT = Path(__file__).resolve().parent.parent / "prisma" / "data" / "users.json"

ROLE_MAP = {
    "truong_ban": "truong_ban",
    "pho_ban": "pho_ban",
    "nhan_vien": "nhan_vien",
    "truong_phong": "truong_phong",
    "pho_phong": "pho_phong",
    "bantgd": "ban_tgd",
    "to_truong": "to_truong",
    "to_pho": "to_pho",
    "admin": "admin",
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["user"] if "user" in wb.sheetnames else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    users = []
    seen = set()
    unknown_roles = {}
    name_from_username = 0

    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        username = clean(col(row, "username"))
        if not username or username.lower() in seen:
            continue
        seen.add(username.lower())

        raw_role = (clean(col(row, "ds_vai_tro")) or "").lower()
        if not raw_role:
            role = "admin" if username.lower() == "admin" else "nhan_vien"
        elif raw_role in ROLE_MAP:
            role = ROLE_MAP[raw_role]
        else:
            role = raw_role
            unknown_roles[raw_role] = unknown_roles.get(raw_role, 0) + 1

        job = clean(col(row, "JobTitles"))
        is_driver = bool(job and job.strip().lower() in ("lái xe", "lai xe"))

        full_name = clean(col(row, "full_name"))
        if not full_name or full_name.strip().lower() == username.lower():
            full_name = full_name or username
            name_from_username += 1

        users.append(
            {
                "username": username,
                "fullName": full_name,
                "dsBan": clean(col(row, "ds_ban")),
                "dsPhong": clean(col(row, "ds_phong")),
                "dsTo": clean(col(row, "ds_to")),
                "role": role,
                "jobTitle": job,
                "email": clean(col(row, "Email")),
                "phone": clean(col(row, "phone")),
                "isDriver": is_driver,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding="utf-8")

    by_role = {}
    for u in users:
        by_role[u["role"]] = by_role.get(u["role"], 0) + 1
    drivers = [u["username"] for u in users if u["isDriver"]]

    print(f"Ghi {len(users)} user -> {OUT}")
    print("Theo vai trò:", json.dumps(by_role, ensure_ascii=False))
    print("Lái xe:", drivers)
    print(f"full_name trùng username (cần bổ sung tên thật): {name_from_username}")
    if unknown_roles:
        print("Vai trò lạ (giữ nguyên):", json.dumps(unknown_roles, ensure_ascii=False))


if __name__ == "__main__":
    main()
